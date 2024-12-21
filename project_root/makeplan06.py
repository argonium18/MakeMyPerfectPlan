import openpyxl
from ortools.sat.python import cp_model
from dataclasses import dataclass, field
from typing import List, Tuple, Optional
from datetime import datetime, timedelta, time
from business_duration import businessDuration
import holidays as pyholidays
from test_jholidays import get_japanese_calendar
from openpyxl.styles import PatternFill, Border, Side
import pandas as pd 

from src.services.convert_to_datetime import convert_to_datetime

@dataclass
class Task:
    id: int
    name: str
    duration: int
    # manually_estimated_start_time: Optional[datetime] = None # 削除予定
    # manually_estimated_end_time: Optional[datetime] = None # 削除予定
    cp_estimated_start_time: Optional[datetime] = None
    cp_estimated_end_time: Optional[datetime] = None
    actual_start_time: Optional[datetime] = None
    actual_end_time: Optional[datetime] = None
    target_end_time: Optional[datetime] = None
    difference: Optional[int] = None  # 差異（実際の終了日時 - 予想終了日時）
    predecessors: List[int] = field(default_factory=list)

class TaskScheduler:
    def __init__(self, num_workers: int, workday_hours: int, start_date: datetime):
        self.tasks: List[Task] = []
        self.num_workers = num_workers
        self.workday_hours = workday_hours
        self.start_date = start_date

    def load_tasks_from_excel(self, file_path: str, sheet_name: str):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
        df = df.dropna(how='all') 

        for _, row in df.iterrows():
            task = Task(
                id=row['タスク番号'],  # 列名が 'id' であると仮定
                name=row['タスク名'],  # 列名が 'name' であると仮定
                duration=row['工数(予想)'],
                cp_estimated_start_time=row['開始日(予想)'],
                cp_estimated_end_time=row['終了日(予想)'],
                predecessors=self._parse_predecessors(row['先行タスク番号'])
            )
            self.tasks.append(task)

    def _parse_predecessors(self, predecessor_str: str) -> List[int]:
        if not predecessor_str or str(predecessor_str).lower() == 'nan':
            return []
        return [int(p) for p in str(predecessor_str).split(',') if p.strip().isdigit()]


    def solve_scheduling(self) -> List[Tuple[int, int, int]]:
        model = cp_model.CpModel()
        start_times = []
        end_times = []
        intervals = []

        for task in self.tasks:
            start = model.NewIntVar(0, 100000, f'start_{task.id}')
            end = model.NewIntVar(0, 100000, f'end_{task.id}')
            interval = model.NewIntervalVar(start, task.duration, end, f'interval_{task.id}')
            start_times.append(start)
            end_times.append(end)
            intervals.append(interval)

        # 依存関係の追加
        for task in self.tasks:
            for predecessor_id in task.predecessors:
                predecessor_index = next((i for i, t in enumerate(self.tasks) if t.id == predecessor_id), None) # 該当の先行タスクが存在しない場合、Noneを設定
                if predecessor_index is not None:
                    model.Add(end_times[predecessor_index] <= start_times[self.tasks.index(task)])
        # リソース制約
        model.AddCumulative(intervals, [1] * len(intervals), self.num_workers)

        # 最速完了時間の最小化
        makespan = model.NewIntVar(0, 100000, 'makespan')
        model.AddMaxEquality(makespan, end_times)
        model.Minimize(makespan)

        # 解の取得
        solver = cp_model.CpSolver()
        status = solver.Solve(model)

        results = []
        if status == cp_model.OPTIMAL:
            for i, task in enumerate(self.tasks):
                start_time = solver.Value(start_times[i])
                end_time = solver.Value(end_times[i])
                
                # CP-SATによる推定開始・終了時間の設定
                task.cp_estimated_start_time, task.cp_estimated_end_time = self.convert_to_datetime(start_time, end_time)
                
                # 手動入力の予想終了時間との差を計算
                if task.target_end_time:
                    task.difference = (task.target_end_time - task.cp_estimated_end_time).total_seconds() / 3600
                results.append((task.id, start_time, end_time))
        
        return results

    def export_results_to_excel(self, file_path: str, sheet_name: str, scheduling_results: List[Tuple[int, int, int]]):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        
        for row in sheet.iter_rows(min_row=2):
            task_id = row[0].value
            if task_id is None:
                break
            

            # 対応するタスクを見つける
            task = next((t for t in self.tasks if t.id == task_id), None)
            if task:
                # D列、E列にCP-SATによる推定開始・終了時間を書き込む
                row[3].value = self.tasks[task_id].cp_estimated_start_time  # D列
                row[4].value = self.tasks[task_id].cp_estimated_end_time    # E列
                # row[9].value = task.difference


                # ガントチャートの図示
                project_start_date = '2024-12-01'
                project_end_date = '2024-12-31'
                year = 2024
                calendar_df = get_japanese_calendar(year,project_start_date,project_end_date)
                transposed_df = calendar_df.T

                # 開始列
                start_col = 10

                # オレンジ色のパターンで塗りつぶし
                orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid') 

                # 罫線の設定
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )

                for r_idx, date in enumerate (calendar_df["date"],start = start_col):
                    if r_idx < len(row):
                        row[r_idx].border = thin_border
                        if task.cp_estimated_start_time.date() <= date.date() <= task.cp_estimated_end_time.date():  # 条件を判定
                            print(date)
                            row[r_idx].fill = orange_fill
 
                # G列に手動入力とCP推定の差を書き込む
                # if task.manually_estimated_end_time is not None:
                #     row[6].value = round(task.difference, 2)
                    
                    # # 時間の乖離に応じて色付け
                    # if task.manual_vs_calculated_end_time_diff > 0:
                    #     row[6].font = openpyxl.styles.Font(color='FF0000')  # 赤色 (CP推定が手動入力より長い)
                    # elif task.manual_vs_calculated_end_time_diff < 0:
                    #     row[6].font = openpyxl.styles.Font(color='0000FF')  # 青色 (CP推定が手動入力より短い)
        
        workbook.save(file_path)

